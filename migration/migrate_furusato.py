# -*- coding: utf-8 -*-
"""家計簿.xlsx の「ふ納税_せ」「ふ納税_あ」→ 家計簿DB 移行スクリプト

furusato_items / furusato_years の2シートだけを bulkImport(replace) するので、
資産・収支など他のデータには一切影響しない。

使い方:
  python migration/migrate_furusato.py                     # JSON生成のみ
  python migration/migrate_furusato.py --upload --url <GAS URL> --token <TOKEN>
"""
import argparse
import json
import sys
import urllib.request
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent
XLSX = BASE.parent / "家計簿.xlsx"
OUT = BASE / "output" / "import_furusato.json"


def num(v):
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    return None


def text(v):
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "-", "#VALUE!", "#N/A", "#REF!"):
        return None
    return s


def year_of(v):
    n = num(v)
    return int(n) if n and 2000 < n < 2100 else None


def clean_name(v):
    s = text(v)
    if not s:
        return None
    return " ".join(s.replace("【楽天市場】", "").split())[:120]


def read_items(ws, person, start_row, cols):
    """cols: (西暦, 商品名, 値段, 自治体, URL, 申請状況, 申請方法, 受取状況) の列番号(1始まり)"""
    items = []
    idx = 1
    for r in range(start_row, ws.max_row + 1):
        cell = lambda c: ws.cell(row=r, column=c).value
        name = clean_name(cell(cols[1]))
        price = num(cell(cols[2]))
        url = text(cell(cols[4]))
        if not name and not price and not url:
            continue
        items.append({
            "id": f"mig-{person}-{idx:03d}",
            "person": person,
            "year": year_of(cell(cols[0])),  # '-' は未購入候補 → 年なし
            "name": name or "（商品名不明）",
            "price": price,
            "municipality": text(cell(cols[3])),
            "url": url,
            "application_status": text(cell(cols[5])) or "未購入",
            "application_method": text(cell(cols[6])),
            "receipt_status": text(cell(cols[7])),
            "memo": None,
        })
        idx += 1
    return items


def read_years_se(ws):
    """ふ納税_せ: E列=対象年 G=上限(想定) H=社保 I=医療費控除 J=税額通知書の年収 K=年収想定"""
    rows = []
    for r in range(4, 20):
        year = year_of(ws.cell(row=r, column=5).value)
        if not year:
            continue
        income = num(ws.cell(row=r, column=10).value) or num(ws.cell(row=r, column=11).value)
        row = {
            "person": "せ",
            "year": year,
            "income": income,
            "social_insurance": num(ws.cell(row=r, column=8).value),
            "medical_deduction": num(ws.cell(row=r, column=9).value),
            "limit_manual": num(ws.cell(row=r, column=7).value),
            "memo": "Excel移行",
        }
        if any(row[k] for k in ("income", "social_insurance", "medical_deduction", "limit_manual")):
            rows.append(row)
    return rows


def read_years_a(ws):
    """ふ納税_あ: J列=対象年 L=上限(想定)"""
    rows = []
    for r in range(3, 20):
        year = year_of(ws.cell(row=r, column=10).value)
        if not year:
            continue
        limit = num(ws.cell(row=r, column=12).value)
        if limit:
            rows.append({
                "person": "あ",
                "year": year,
                "income": None,
                "social_insurance": None,
                "medical_deduction": None,
                "limit_manual": limit,
                "memo": "Excel移行",
            })
    return rows


def build_payload():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws_se = wb["ふ納税_せ"]
    ws_a = wb["ふ納税_あ"]
    items = read_items(ws_se, "せ", 19, (1, 2, 4, 5, 6, 7, 8, 9)) + read_items(ws_a, "あ", 2, (1, 2, 3, 4, 5, 6, 7, 8))
    years = read_years_se(ws_se) + read_years_a(ws_a)
    return {"action": "bulkImport", "mode": "replace", "furusato_items": items, "furusato_years": years}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--upload", action="store_true")
    ap.add_argument("--url")
    ap.add_argument("--token")
    args = ap.parse_args()

    p = build_payload()
    OUT.parent.mkdir(exist_ok=True)
    OUT.write_text(json.dumps(p, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"生成: {OUT}\n")
    for person in ("せ", "あ"):
        items = [i for i in p["furusato_items"] if i["person"] == person]
        years = [y for y in p["furusato_years"] if y["person"] == person]
        total = {}
        for i in items:
            if i["year"] and i["price"] and i["application_status"] != "未購入":
                total[i["year"]] = total.get(i["year"], 0) + i["price"]
        print(f"[{person}] items: {len(items)}件 / years: {len(years)}件 / 年別購入合計: " +
              ", ".join(f"{y}={int(v):,}円" for y, v in sorted(total.items())))

    if args.upload:
        if not args.url or not args.token:
            ap.error("--upload には --url と --token が必要です")
        payload = dict(p, token=args.token)
        req = urllib.request.Request(
            args.url,
            data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            headers={"Content-Type": "text/plain;charset=utf-8"},
            method="POST",
        )
        with urllib.request.urlopen(req) as res:
            body = json.loads(res.read().decode("utf-8"))
        if not body.get("ok"):
            print(f"投入失敗: {body.get('error')}", file=sys.stderr)
            sys.exit(1)
        print("投入成功:", json.dumps(body["data"].get("imported") if isinstance(body["data"], dict) else body["data"], ensure_ascii=False)[:200])


if __name__ == "__main__":
    main()
