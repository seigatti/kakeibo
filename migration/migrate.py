# -*- coding: utf-8 -*-
"""家計簿.xlsx → 家計簿DB(Googleシート) 移行スクリプト

使い方:
  1. JSON生成のみ（内容確認用）:
       python migrate.py
     → migration/output/import.json が生成される

  2. GAS API へ投入（GASデプロイ後）:
       python migrate.py --upload --url <GASウェブアプリURL> --token <APIトークン>

投入は bulkImport(replace) なので、何度実行してもシートは同じ状態になる。
"""
import argparse
import json
import sys
import urllib.request
from datetime import date, datetime
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent
XLSX = BASE.parent / "家計簿.xlsx"
OUT = BASE / "output" / "import.json"

VARIABLE_CATEGORIES = ["ガス代", "電気代", "上下水道", "ガソリン代", "高速料金", "ケータイ料金"]


def ymd(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return None


def ym(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m")
    return None


def num(v):
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    if isinstance(v, str):
        try:
            return round(float(v.replace(",", "")), 2)
        except ValueError:
            return None
    return None


def read_assets(wb):
    """★資産額合計まとめ★: id | 年月日 | 投資 | 現金 | 年金 | 合計"""
    ws = wb["★資産額合計まとめ★"]
    rows = []
    for r in ws.iter_rows(min_row=2, max_col=5):
        d = ymd(r[1].value)
        inv, cash, pension = num(r[2].value), num(r[3].value), num(r[4].value)
        if d and (inv is not None or cash is not None):
            rows.append({
                "date": d,
                "investment": inv,
                "cash": cash,
                "pension": pension,
                "mf_profit": None,
                "memo": "Excel移行",
            })
    return rows


def read_expenses(wb):
    """★固定費集計 左表: 引き落とし月 × 変動費6カテゴリ（B〜G列）"""
    ws = wb["★固定費集計"]
    rows = []
    for r in ws.iter_rows(min_row=5, max_col=7):
        month = ym(r[0].value)
        if not month:
            continue
        vals = [num(c.value) for c in r[1:7]]
        if all(v is None for v in vals):
            continue  # 実績のない将来行（予測式のみ）はスキップ
        for cat, v in zip(VARIABLE_CATEGORIES, vals):
            if v is not None:
                rows.append({"month": month, "category": cat, "amount": v})
    return rows


def read_fixed_costs(wb):
    """★固定費集計 右表(J〜N列): 固定費用 | 引き落とし | 頻度 | 金額 | 金額(新)"""
    ws = wb["★固定費集計"]
    freq_map = [("2年", "2年"), ("年", "年"), ("月", "月")]
    rows = []
    idx = 1
    for r in ws.iter_rows(min_row=6, max_row=40, min_col=10, max_col=14):
        name = r[0].value
        if not name or str(name).strip() == "毎月固定額":
            continue
        freq_raw = str(r[2].value or "")
        freq = next((v for k, v in freq_map if k in freq_raw), None)
        if not freq:
            continue
        amount = num(r[4].value) or num(r[3].value)
        if amount is None:
            continue
        rows.append({
            "id": f"mig-{idx:02d}",
            "name": str(name).strip(),
            "amount": amount,
            "frequency": freq,
            "start_month": None,
            "end_month": None,
            "memo": (str(r[1].value).strip() if r[1].value else None),
        })
        idx += 1
    return rows


def read_zaim_net(wb):
    """★非運用資産収支: 月 | 金額（Zaim月次収支の転記）"""
    ws = wb["★非運用資産収支"]
    rows = []
    for r in ws.iter_rows(min_row=1, max_col=2):
        month = ym(r[0].value)
        v = num(r[1].value)
        if month and v is not None:
            rows.append({"month": month, "amount": v})
    return rows


def build_payload():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    payload = {
        "action": "bulkImport",
        "mode": "replace",
        "assets": read_assets(wb),
        "expenses": read_expenses(wb),
        "fixed_costs": read_fixed_costs(wb),
        "zaim_net": read_zaim_net(wb),
        "settings": [{"key": "expense_categories", "value": ",".join(VARIABLE_CATEGORIES)}],
    }
    return payload


def summarize(p):
    a = p["assets"]
    print(f"assets      : {len(a)} 件 ({a[0]['date']} 〜 {a[-1]['date']})")
    latest = a[-1]
    total = sum(v for v in [latest["investment"], latest["cash"], latest["pension"]] if v)
    print(f"  最新スナップショット合計: {total:,.0f} 円 (Excelの合計と一致するか確認)")
    months = sorted({e['month'] for e in p['expenses']})
    print(f"expenses    : {len(p['expenses'])} 件 ({months[0]} 〜 {months[-1]}, {len(months)} ヶ月)")
    fc = p["fixed_costs"]
    monthly = sum(f["amount"] / {"月": 1, "年": 12, "2年": 24}[f["frequency"]] for f in fc)
    print(f"fixed_costs : {len(fc)} 件 / 月割り合計 {monthly:,.0f} 円")
    print("  ※Excelの「毎月固定額34,250円」行は家賃等が0円だった頃の古い式。")
    print("    Excel合計列の固定分109,564円 = 全項目112,734円 − ジム3,170円 と整合。")
    print(f"zaim_net    : {len(p['zaim_net'])} 件")


def upload(payload, url, token):
    payload = dict(payload, token=token)
    req = urllib.request.Request(
        url,
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={"Content-Type": "text/plain;charset=utf-8"},
        method="POST",
    )
    with urllib.request.urlopen(req) as res:
        body = json.loads(res.read().decode("utf-8"))
    if not body.get("ok"):
        print(f"投入失敗: {body.get('error')}", file=sys.stderr)
        sys.exit(1)
    print("投入成功:", json.dumps(body["data"], ensure_ascii=False))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--upload", action="store_true", help="GAS API へ投入する")
    ap.add_argument("--url", help="GAS ウェブアプリURL (https://script.google.com/macros/s/.../exec)")
    ap.add_argument("--token", help="APIトークン")
    args = ap.parse_args()

    payload = build_payload()
    OUT.parent.mkdir(exist_ok=True)
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"生成: {OUT}\n")
    summarize(payload)

    if args.upload:
        if not args.url or not args.token:
            ap.error("--upload には --url と --token が必要です")
        upload(payload, args.url, args.token)


if __name__ == "__main__":
    main()
